import os
import re
import time

from openpyxl import load_workbook
from openpyxl.worksheet .worksheet import Worksheet
from openpyxl.styles import PatternFill
import json


def set_passed(sheet: Worksheet,
               row: int,
               column: int,
               passed: bool = None):
    current_val = sheet.cell(row=row, column=column).value

    if passed is None or current_val == "NA":
        sheet.cell(row=row, column=column, value="NA")
        return

    if current_val == "No":
        return

    sheet.cell(row=row, column=column, value="Yes" if passed else "No")


def check_overall_passed(sheet: Worksheet, row: int, columns: list[int]) -> bool:
    """Check if the overall passed status is 'Yes'."""
    for column in columns:
        cell_value = sheet.cell(row=row, column=column).value
        if cell_value != "Yes":
            return False

    return True


class SampleReturns:
    def __init__(self, filename: str):
        self._filename = filename
        self._create_backup()

        self._wb = load_workbook(filename)
        self._saved = False

    def __del__(self):
        if hasattr(self, '_wb') and self._wb is not None:
            if not self._saved:
                print(f"Warning: {self._filename} was not saved before closing.")

            self._wb.close()

    def _create_backup(self):
        """Create a backup of the current workbook."""
        backup_filename = self._filename[:-5] + "_backup.xlsx"
        load_workbook(self._filename).save(backup_filename)
        print(f"Backup created: {backup_filename}")

    def save(self):
        self._wb.active_cell = self._wb["HW"]
        self._wb.save(self._filename)
        self._saved = True

    def _get_row_by_serial_number(self, serial_number: str) -> int:
        """Get the row number for a given serial number."""
        sheet = self._wb["HW"]

        for row in range(2, sheet.max_row + 1):
            current_serial_nr = sheet.cell(row=row, column=7).value
            if current_serial_nr == "":
                return -1  # Serial number not found
            elif current_serial_nr == serial_number:
                return row

        return -1

    def get_row_by_sample_label(self, sheet: Worksheet, sample_label: str) -> int:
        """Get the row number for a given serial number."""
        for row in range(1, sheet.max_row + 1):
            current_serial_nr = sheet.cell(row=row, column=1).value

            if current_serial_nr is None:
                continue

            if current_serial_nr.startswith("="):
                index_tuple = current_serial_nr[1:].split("!")
                current_serial_nr = self._wb[index_tuple[0]][index_tuple[1]].value

            if current_serial_nr == "":
                return -1  # Serial number not found
            elif current_serial_nr == sample_label:
                return row

        return -1

    def _replace_none_with_na(self, sample_label: str):
        """Replace None values with 'NA' in the specified sample row."""
        skip = ["UWB RX Sensitivity", "UWB Ranging", "UWB AoA"]

        sheets = self._wb.sheetnames
        for sheet_name in sheets:
            if sheet_name in skip or not sheet_name.startswith("UWB"):
                continue

            sheet = self._wb[sheet_name]
            row = self.get_row_by_sample_label(sheet, sample_label)

            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is None:
                    sheet.cell(row=row, column=col, value="NA")
                    print(f"Replaced None with 'NA' in {sheet_name} at row {row}, column {col}")


    def _process_frequency_config(self, measurements: dict, sample_label: str):
        sheet = self._wb["UWB Frequency"]
        row = self.get_row_by_sample_label(sheet=sheet, sample_label=sample_label)

        hmm_value = measurements.get("calibration_value_itac")
        if hmm_value is None:
            hmm_value = (None, None)
        sheet.cell(row=row, column=3, value=hmm_value[1])
        sheet.cell(row=row, column=4, value=hmm_value[0])

        sheet.cell(row=row, column=5, value=measurements["target_value"])
        sheet.cell(row=row, column=6, value=measurements.get("itac_measurement"))
        sheet.cell(row=row, column=7, value=measurements["measured_value"])

        passed = measurements.get("passed", None)
        set_passed(sheet=sheet, row=row, column=8, passed=passed)

    def _process_cw_config(self, measurements: dict, sample_label: str):
        antenna_config = measurements["antenna_config"]
        antenna_idx = re.search(r'\d+', antenna_config).group(0)

        sheet = self._wb[f"UWB Antenna {antenna_idx}"]
        row = self.get_row_by_sample_label(sheet=sheet, sample_label=sample_label)

        channel = measurements["channel"]
        if channel == "UWB_CH8":
            start_col = 4
        elif channel == "UWB_CH9":
            start_col = 12
        else:
            raise ValueError(f"Unknown channel: {channel}")

        hmm_value = measurements.get("calibration_value_itac")
        if hmm_value is not None:
            sheet.cell(row=row, column=start_col, value=hmm_value)

        sheet.cell(row=row, column=start_col + 1, value=measurements["target_value"])
        sheet.cell(row=row, column=start_col + 2, value=measurements.get("itac_measurement"))
        sheet.cell(row=row, column=start_col + 3, value=measurements["measured_value"])

        passed = measurements.get("passed", None)
        set_passed(sheet=sheet, row=row, column=start_col + 6, passed=passed)

    def _process_sts_config(self, measurements: dict, sample_label: str):
        antenna_config = measurements["antenna_config"]
        antenna_idx = re.search(r'\d+', antenna_config).group(0)

        sheet = self._wb[f"UWB Antenna {antenna_idx}"]
        row = self.get_row_by_sample_label(sheet=sheet, sample_label=sample_label)

        channel = measurements["channel"]
        if channel == "UWB_CH8":
            start_col = 4
        elif channel == "UWB_CH9":
            start_col = 12
        else:
            raise ValueError(f"Unknown channel: {channel}")

        hmm_value = measurements.get("calibration_value_itac")
        if hmm_value is not None:
            sheet.cell(row=row, column=start_col, value=hmm_value)

        sheet.cell(row=row, column=start_col+4, value=measurements["target_value"])
        sheet.cell(row=row, column=start_col+5, value=measurements["measured_value"])

        passed = measurements.get("passed", None)
        set_passed(sheet=sheet, row=row, column=start_col+6, passed=passed)

    def _process_radar_config(self, measurements: dict, sample_label: str):
        sheet = self._wb[f"UWB Antenna 1 (Radar)"]
        row = self.get_row_by_sample_label(sheet=sheet, sample_label=sample_label)

        channel = measurements["channel"]
        if channel == "UWB_CH8":
            start_col = 6
        elif channel == "UWB_CH9":
            start_col = 11
        else:
            raise ValueError(f"Unknown channel: {channel}")

        hmm_value = measurements.get("calibration_value_itac")
        sheet.cell(row=row, column=3, value=hmm_value)

        sheet.cell(row=row, column=start_col, value=measurements["target_value"])
        sheet.cell(row=row, column=start_col+1, value=measurements["measured_value"])

        passed = measurements.get("passed", None)
        set_passed(sheet=sheet, row=row, column=start_col+2, passed=passed)

    def upload_json(self, filename: str):
        """Load a JSON file and return its content."""
        with open(filename, 'r') as file:
            measurement_data = json.load(file)

        row = self._get_row_by_serial_number(measurement_data["serial_number"])
        if row == -1:
            print(f"Serial number {measurement_data['serial_number']} not found in the workbook.")
            return

        sample_label = self._wb["HW"].cell(row=row, column=1).value
        self._wb["HW"].cell(row=row, column=15, value=measurement_data["variant_type"])

        eol = measurement_data.get("eol", "NA")
        if eol != "NA" and isinstance(eol, int):
            eol = "EOL" + str(eol)
        self._wb["HW"].cell(row=row, column=2, value=eol)

        is_locked = measurement_data.get("is_locked", "NA")
        if is_locked != "NA":
            is_locked = "Yes" if is_locked else "No"
        self._wb["HW"].cell(row=row, column=3, value=is_locked)
        self._wb["HW"].cell(row=row, column=4, value=measurement_data.get("eol_date", "NA"))
        self._wb["HW"].cell(row=row, column=5, value=measurement_data.get("measurement_date", "NA"))


        for config in measurement_data["antenna_configs"]:
            calibration_type = config["calibration_type"].lower()

            if "frequency" in calibration_type:
                self._process_frequency_config(measurements=config, sample_label=sample_label)
            elif "cw" in calibration_type:
                self._process_cw_config(measurements=config, sample_label=sample_label)
            elif "sts" in calibration_type:
                self._process_sts_config(measurements=config, sample_label=sample_label)
            elif "radar" in calibration_type:
                self._process_radar_config(measurements=config, sample_label=sample_label)

        self._replace_none_with_na(sample_label=sample_label)

def resync_json():
    working_dir = os.path.abspath(".")

    excel_interface = SampleReturns(
        filename=os.path.join(working_dir,
                              "excel_files",
                              "2024-12-19_BMW_FBD6_Reference_Sample_Docu_UWB+BLE_Final.xlsx"))

    files = os.listdir(os.path.join(working_dir, "export", "measurement_data"))
    for file in files:
        if not file.endswith(".json"):
            continue

        filepath = os.path.join(working_dir, "export", "measurement_data", file)
        excel_interface.upload_json(filepath)


    excel_interface.save()



if __name__ == '__main__':
    resync_json()